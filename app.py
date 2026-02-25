import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import gspread
import openpyxl
import datetime
import calendar
import yfinance as yf
import requests
import re
import io
import json
import os
import numpy as np
from PIL import Image

try:
    import google.generativeai as genai
    HAS_GENAI = True
except ImportError:
    HAS_GENAI = False

try:
    import pdfplumber
    HAS_PDF = True
except ImportError:
    HAS_PDF = False

# ==========================================
# 1. PAGE SETUP & THEME ENGINE
# ==========================================
st.set_page_config(page_title="Jaynik's Finance Dashboard", page_icon="💎", layout="wide")

if 'theme' not in st.session_state:
    st.session_state['theme'] = 'Light'

# --- CSS THEMES ---
DARK_THEME = """
<style>
    /* 1. GLOBAL BACKGROUNDS */
    [data-testid="stAppViewContainer"], [data-testid="stHeader"], .main {
        background-color: #000000 !important;
    }
    /* 2. SIDEBAR */
    [data-testid="stSidebar"], [data-testid="stSidebar"] > div:first-child {
        background-color: #050505 !important; border-right: 1px solid #222 !important;
    }
    /* 3. TEXT GLOW */
    html, body, p, label, span, li, td, th, div {
        color: #ffffff !important; text-shadow: 0 0 1px rgba(255, 255, 255, 0.4) !important;
    }
    /* 4. INPUTS */
    [data-baseweb="select"], .stSelectbox div { background-color: #111 !important; color: #fff !important; }
    /* 5. KPI GLOWS */
    .glow-income { border-top: 4px solid #00c853 !important; background: #0a0a0a !important; box-shadow: 0 4px 15px rgba(0, 200, 83, 0.2) !important; border-radius: 10px; padding: 15px; margin-bottom: 20px; }
    .glow-expenses { border-top: 4px solid #ff1744 !important; background: #0a0a0a !important; box-shadow: 0 4px 15px rgba(255, 23, 68, 0.2) !important; border-radius: 10px; padding: 15px; margin-bottom: 20px; }
    .glow-savings { border-top: 4px solid #2979ff !important; background: #0a0a0a !important; box-shadow: 0 4px 15px rgba(41, 121, 255, 0.2) !important; border-radius: 10px; padding: 15px; margin-bottom: 20px; }
    .glow-balance { border-top: 4px solid #ffd600 !important; background: #0a0a0a !important; box-shadow: 0 4px 15px rgba(255, 214, 0, 0.2) !important; border-radius: 10px; padding: 15px; margin-bottom: 20px; }
    .kpi-value { font-size: 2rem; font-weight: 800; color: #fff; }
    .kpi-title { font-size: 0.9rem; font-weight: 700; text-transform: uppercase; letter-spacing: 1px; margin-bottom: 5px; }
</style>
"""
LIGHT_THEME = """
<style>
    [data-testid="stAppViewContainer"] { background-color: #f8f9fa; color: #212529; }
    .glow-card { background: #ffffff; padding: 20px; border-radius: 12px; box-shadow: 0 4px 12px rgba(0,0,0,0.05); }
</style>
"""

# Standard Financial Year List
FY_MONTHS = ['April', 'May', 'June', 'July', 'August', 'September', 'October', 'November', 'December', 'January', 'February', 'March']

# --- HELPER FUNCTIONS (REQUIRED) ---
def format_inr(number):
    try:
        n = float(number); is_neg = n < 0; n = abs(n)
        s, *d = "{:.0f}".format(n).partition(".")
        r = ",".join([s[x-2:x] for x in range(-3, -len(s), -2)][::-1] + [s[-3:]])
        return "-" + "".join([r] + d) if is_neg else "".join([r] + d)
    except: return str(number)

def get_financial_year(date):
    if pd.isna(date): return "Unknown"
    return f"FY {date.year}-{str(date.year+1)[-2:]}" if date.month >= 4 else f"FY {date.year-1}-{str(date.year)[-2:]}"
def init_connection():
    scope = ['https://spreadsheets.google.com/feeds', 'https://www.googleapis.com/auth/drive']
    
    # CASE 1: Local Laptop (Uses file)
    if os.path.exists('credentials.json'):
        creds = ServiceAccountCredentials.from_json_keyfile_name('credentials.json', scope)
    
    # CASE 2: Streamlit Cloud (Uses Secrets Dictionary)
    else:
        # Create a dictionary directly from the secrets
        # This bypasses the strict JSON parser entirely
        creds_dict = {
            "type": st.secrets["gcp_service_account"]["type"],
            "project_id": st.secrets["gcp_service_account"]["project_id"],
            "private_key_id": st.secrets["gcp_service_account"]["private_key_id"],
            "private_key": st.secrets["gcp_service_account"]["private_key"],
            "client_email": st.secrets["gcp_service_account"]["client_email"],
            "client_id": st.secrets["gcp_service_account"]["client_id"],
            "auth_uri": st.secrets["gcp_service_account"]["auth_uri"],
            "token_uri": st.secrets["gcp_service_account"]["token_uri"],
            "auth_provider_x509_cert_url": st.secrets["gcp_service_account"]["auth_provider_x509_cert_url"],
            "client_x509_cert_url": st.secrets["gcp_service_account"]["client_x509_cert_url"]
        }
        creds = ServiceAccountCredentials.from_json_keyfile_dict(creds_dict, scope)
        
    client = gspread.authorize(creds)
    return client

# --- PART B: THE DATA LOADER (Uses the connection above) ---
@st.cache_data(ttl=5)
def load_data():
    try:
        client = init_connection()  # <--- This calls the function above
        sh = client.open("Finance Tracker")
    except Exception as e:
        st.error(f"🚨 Connection Error: {e}")
        st.stop()

    # Helper to safely read a tab
    def get_df(worksheet_name):
        try:
            ws = sh.worksheet(worksheet_name)
            data = ws.get_all_records()
            return pd.DataFrame(data)
        except: return pd.DataFrame()

    # Load Main Data
    df = get_df('Budget Tracking')
    
    # Clean Data
    if not df.empty and 'Date' in df.columns:
        df['Date'] = pd.to_datetime(df['Date'], errors='coerce')
        df = df.dropna(subset=['Date'])
        df['Year'] = df['Date'].dt.year 
        df['FY'] = df['Date'].apply(get_financial_year)
        df['Month'] = df['Date'].dt.month_name()
        df['Amount'] = pd.to_numeric(df['Amount'], errors='coerce').fillna(0)

    # Load Budget
    budget_raw = get_df('Budget Planning')
    budget_melted['Amount'] = pd.to_numeric(budget_melted['Amount'], errors='coerce').fillna(0)

    # 1. RETURN THE DATA (Indented so it is part of the function)
    return df, budget_melted, budget_raw, get_df('Credit Cards'), get_df('Loans'), get_df('Physical Assets'), get_df('Splitwise'), get_df('Subscriptions'), get_df('Goals')

# 2. CALL THE FUNCTION (NO Indentation - This runs the code)
df, budget_df, budget_raw_df, cc_df, loan_df, assets_df, split_df, subs_df, goals_df = load_data()

# ---------------------------------------------------------
# NOW your split logic will work because split_df exists
split_users = set(["Partner"])
if not split_df.empty:
    split_users.update(split_df['Payer'].dropna().astype(str).unique())
    split_users.update(split_df['Debtor'].dropna().astype(str).unique())
if "Jaynik" in split_users: split_users.remove("Jaynik")
split_users = sorted(list(split_users))

# ==========================================
# 3. GLOBAL DATA FETCHERS
# ==========================================
@st.cache_data(ttl=3600) 
def fetch_amfi_data():
    amfi_dict = {}; mf_dropdown_list = []
    try:
        url = "https://www.amfiindia.com/spages/NAVAll.txt"
        response = requests.get(url, timeout=10)
        for line in response.text.split('\n'):
            parts = line.split(';')
            if len(parts) >= 5 and parts[0].isdigit():
                code = parts[0]; name = parts[3]
                try: amfi_dict[code] = {'name': name, 'nav': float(parts[4])}; mf_dropdown_list.append(f"{name} [{code}]")
                except: pass
    except: pass
    mf_dropdown_list.sort(); return amfi_dict, mf_dropdown_list

@st.cache_data(ttl=86400) 
def fetch_nse_data():
    nse_dict = {}; nse_dropdown_list = []
    try:
        url = "https://nsearchives.nseindia.com/content/equities/EQUITY_L.csv"
        headers = {'User-Agent': 'Mozilla/5.0'}
        response = requests.get(url, headers=headers, timeout=10)
        if response.status_code == 200:
            df_nse = pd.read_csv(io.StringIO(response.text))
            for _, row in df_nse.iterrows():
                symbol = str(row['SYMBOL']).strip(); name = str(row['NAME OF COMPANY']).strip().title()
                nse_dict[symbol] = name; nse_dropdown_list.append(f"{name} [{symbol}]")
    except: pass
    etf_list = {"NIFTYBEES": "Nippon India Nifty 50 ETF", "BANKBEES": "Nippon India Bank ETF"}
    for sym, name in etf_list.items(): nse_dict[sym] = name; nse_dropdown_list.append(f"⭐ {name} [{sym}]") 
    nse_dropdown_list.sort(); return nse_dict, nse_dropdown_list

@st.cache_data(ttl=300) 
def get_market_data(tickers, amfi_dict, nse_dict):
    prices = {}; names = {}
    for t in tickers:
        t_str = str(t)
        if t_str in amfi_dict: prices[t_str] = amfi_dict[t_str]['nav']; names[t_str] = amfi_dict[t_str]['name']
        else:
            clean_ticker = t_str.replace('.NS', '').replace('.BO', '')
            if clean_ticker in nse_dict: names[t_str] = nse_dict[clean_ticker]
            else: names[t_str] = t_str 
            try:
                search_ticker = t_str if "." in t_str or "-" in t_str or t_str.isdigit() else f"{t_str}.NS"
                hist = yf.Ticker(search_ticker).history(period="1d")
                prices[t_str] = float(hist['Close'].iloc[-1]) if not hist.empty else 0.0
            except: prices[t_str] = 0.0
    return prices, names

# --- RESTORED FUNCTION ---
@st.cache_data(ttl=86400)
def fetch_benchmark_history(start_date):
    try:
        start_date_str = pd.to_datetime(start_date).strftime('%Y-%m-%d')
        nifty = yf.Ticker("^NSEI").history(start=start_date_str)
        nifty.index = nifty.index.tz_localize(None).normalize()
        idx = pd.date_range(start=nifty.index.min(), end=pd.Timestamp.today().normalize())
        nifty = nifty.reindex(idx).ffill().bfill()
        return nifty['Close']
    except: return pd.Series()

amfi_data_dict, amfi_dropdown = fetch_amfi_data()
nse_data_dict, nse_dropdown = fetch_nse_data()

# ==========================================
# 4. SIDEBAR: NAVIGATION
# ==========================================
st.sidebar.markdown("## 🎨 App Theme")
theme_choice = st.sidebar.radio("Select Mode:", ["Light", "Dark Mode"], horizontal=True)

if theme_choice == "Dark Mode":
    st.markdown(DARK_THEME, unsafe_allow_html=True); chart_text_color = "#e0e0e0"; root_node_color = "#333333"
else:
    st.markdown(LIGHT_THEME, unsafe_allow_html=True); chart_text_color = "#212529"; root_node_color = "#e9ecef"

current_user = "Jaynik"

st.sidebar.markdown("---")
st.sidebar.markdown("## 🧭 Navigation")
page = st.sidebar.radio("Go To Screen", ["🏠 Main Dashboard (I&E)", "💰 Budget Planner", "📈 Investment Tracker", "💳 Credit Cards", "🔄 Subscription Radar", "🤝 Splitwise / Settles", "⚖️ Net Worth & Goals", "📸 AI Bill Scanner", "📝 Transactions"])

st.sidebar.markdown("---")
st.sidebar.header("📅 Financial Year Filters")
data_fys = df['FY'].dropna().unique().tolist()
current_fy = get_financial_year(datetime.datetime.today())
if current_fy not in data_fys: data_fys.append(current_fy)
data_fys = sorted(list(set(data_fys)), reverse=True)
selected_fy = st.sidebar.selectbox("Select Financial Year", ["All Years"] + data_fys)
all_months = list(calendar.month_name)[1:]
selected_month = st.sidebar.selectbox("Select Month", ["All Months"] + all_months)

filtered_df = df.copy(); filtered_budget_df = budget_df.copy()
if selected_fy != "All Years":
    filtered_df = filtered_df[filtered_df['FY'] == selected_fy]
if selected_month != "All Months":
    filtered_df = filtered_df[filtered_df['Month'] == selected_month]
    filtered_budget_df = filtered_budget_df[filtered_budget_df['Month'] == selected_month]

def process_pie_data(target_df, threshold=0.05):
    if target_df.empty: return target_df
    agg_df = target_df.groupby('Category')['Amount'].sum().reset_index()
    total = agg_df['Amount'].sum()
    if total == 0: return agg_df
    agg_df['Percent'] = agg_df['Amount'] / total
    main_cats = agg_df[agg_df['Percent'] >= threshold].copy()
    small_cats = agg_df[agg_df['Percent'] < threshold]
    if not small_cats.empty:
        others_row = pd.DataFrame([{'Category': 'Others', 'Amount': small_cats['Amount'].sum(), 'Percent': small_cats['Percent'].sum()}])
        main_cats = pd.concat([main_cats, others_row], ignore_index=True)
    return main_cats

# ==========================================
# 5. PAGE LOGIC MANAGER
# ==========================================

if page == "🏠 Main Dashboard (I&E)":
    if current_user != "Household (Combined)":
        pending_alerts = split_df[(split_df['Debtor'] == current_user) & (split_df['Status'] == 'Pending')]
        if not pending_alerts.empty:
            total_owed_by_me = pending_alerts['Split Amount'].sum()
            st.markdown(f"<div class='alert-box'><b>🔔 You have pending Splitwise settlements!</b> You owe ₹{total_owed_by_me:,.0f} to other members.</div>", unsafe_allow_html=True)
            
    if filtered_df.empty: st.warning("⚠️ No transactions found for the selected Financial Year/Month.")

    total_income = filtered_df[filtered_df['Type'] == 'Income']['Amount'].sum()
    total_expenses = filtered_df[filtered_df['Type'] == 'Expenses']['Amount'].sum()
    total_savings = filtered_df[filtered_df['Type'] == 'Savings']['Amount'].sum()
    net_balance = total_income - total_expenses - total_savings

    col1, col2, col3, col4 = st.columns(4)
    with col1: st.markdown(f"<div class='glow-card glow-green'><div class='kpi-title'>TOTAL INCOME</div><div class='kpi-value' style='color:#00c853'>₹{total_income:,.0f}</div></div>", unsafe_allow_html=True)
    with col2: st.markdown(f"<div class='glow-card glow-red'><div class='kpi-title'>TOTAL EXPENSES</div><div class='kpi-value' style='color:#d50000'>₹{total_expenses:,.0f}</div></div>", unsafe_allow_html=True)
    with col3: st.markdown(f"<div class='glow-card glow-blue'><div class='kpi-title'>TOTAL SAVINGS</div><div class='kpi-value' style='color:#2962ff'>₹{total_savings:,.0f}</div></div>", unsafe_allow_html=True)
    with col4: st.markdown(f"<div class='glow-card glow-gold'><div class='kpi-title'>NET BALANCE</div><div class='kpi-value' style='color:#ffd600'>₹{net_balance:,.0f}</div></div>", unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)
    col_pie1, col_pie2, col_pie3 = st.columns(3)
    with col_pie1:
        st.markdown("<div class='chart-box'><h4 style='text-align: center;'>💼 INCOME</h4>", unsafe_allow_html=True)
        income_df = filtered_df[(filtered_df['Type'] == 'Income') & (filtered_df['Amount'] > 0)]
        income_df = process_pie_data(income_df, 0.05)
        if not income_df.empty:
            fig_inc = px.pie(income_df, values='Amount', names='Category', hole=0.5, color_discrete_sequence=px.colors.sequential.Teal)
            fig_inc.update_layout(paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)', font=dict(color=chart_text_color), margin=dict(t=0, b=20, l=0, r=0), showlegend=False)
            fig_inc.update_traces(textposition='inside', textinfo='percent+label')
            st.plotly_chart(fig_inc, use_container_width=True)
        st.markdown("</div>", unsafe_allow_html=True)

    with col_pie2:
        st.markdown("<div class='chart-box'><h4 style='text-align: center;'>📉 EXPENSES</h4>", unsafe_allow_html=True)
        expense_df = filtered_df[(filtered_df['Type'] == 'Expenses') & (filtered_df['Amount'] > 0)]
        expense_df = process_pie_data(expense_df, 0.05)
        if not expense_df.empty:
            fig_exp = px.pie(expense_df, values='Amount', names='Category', hole=0.5, color_discrete_sequence=px.colors.sequential.Reds)
            fig_exp.update_layout(paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)', font=dict(color=chart_text_color), margin=dict(t=0, b=20, l=0, r=0), showlegend=False)
            fig_exp.update_traces(textposition='inside', textinfo='percent+label')
            st.plotly_chart(fig_exp, use_container_width=True)
        st.markdown("</div>", unsafe_allow_html=True)

    with col_pie3:
        st.markdown("<div class='chart-box'><h4 style='text-align: center;'>🏦 SAVINGS</h4>", unsafe_allow_html=True)
        savings_df = filtered_df[(filtered_df['Type'] == 'Savings') & (filtered_df['Amount'] > 0)]
        savings_df = process_pie_data(savings_df, 0.05)
        if not savings_df.empty:
            fig_sav = px.pie(savings_df, values='Amount', names='Category', hole=0.5, color_discrete_sequence=px.colors.sequential.Blues)
            fig_sav.update_layout(paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)', font=dict(color=chart_text_color), margin=dict(t=0, b=20, l=0, r=0), showlegend=False)
            fig_sav.update_traces(textposition='inside', textinfo='percent+label')
            st.plotly_chart(fig_sav, use_container_width=True)
        st.markdown("</div>", unsafe_allow_html=True)

    st.markdown("### 📊 Actuals vs. Budget")
    if not filtered_budget_df.empty:
        # Aggregating actuals
        monthly_actuals = filtered_df.groupby(['Month', 'Type'])['Amount'].sum().reset_index()
        monthly_actuals['Source'] = 'Actual'
        
        # Aggregating budget (Already melted)
        monthly_budget = filtered_budget_df.groupby(['Month', 'Type'])['Amount'].sum().reset_index()
        monthly_budget['Source'] = 'Budget'
        
        combined_df = pd.concat([monthly_actuals, monthly_budget])
        
        # Sorting Months Properly using Global FY_MONTHS
        combined_df['Month'] = pd.Categorical(combined_df['Month'], categories=FY_MONTHS, ordered=True)
        combined_df = combined_df.sort_values('Month')
        
        combined_df['Type_Source'] = combined_df['Type'] + ' (' + combined_df['Source'] + ')'
        color_map = {'Income (Actual)': '#00c853', 'Income (Budget)': '#b9f6ca', 'Expenses (Actual)': '#d50000', 'Expenses (Budget)': '#ff8a80', 'Savings (Actual)': '#2962ff', 'Savings (Budget)': '#82b1ff'}
        
        # Enforce X Axis Order specifically for Indian FY
        fig_bar = px.bar(combined_df, x='Month', y='Amount', color='Type_Source', barmode='group', color_discrete_map=color_map, category_orders={"Month": FY_MONTHS})
        fig_bar.update_layout(paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)', font=dict(color=chart_text_color), xaxis=dict(showgrid=False), yaxis=dict(showgrid=True, gridcolor='rgba(128,128,128,0.2)'), legend=dict(orientation="h", y=1.1, title=None))
        st.plotly_chart(fig_bar, use_container_width=True)
    else: st.info("No budget data set. Go to 'Budget Planner' to set goals.")

elif page == "💰 Budget Planner":
    st.markdown("<h2>💰 Budget Planner</h2>", unsafe_allow_html=True)
    st.write("Manage your monthly financial limits. Use the tabs below to set new budgets or view the full matrix.")
    
    # NEW TABS FOR BUDGET VIEW
    tab_input, tab_month_view, tab_full_matrix = st.tabs(["📝 Input & Update", "🗓️ Monthly View", "📊 Full Matrix"])
    
    with tab_input:
        st.write("### Set a Budget Goal")
        # NOTE: Removed st.form to allow instant update of Month dropdown
        c1, c2 = st.columns(2)
        with c1:
            b_type = st.selectbox("Type", ["Income", "Expenses", "Savings"])
            existing_cats = df[df['Type'] == b_type]['Category'].unique().tolist()
            b_cat = st.selectbox("Category", existing_cats + ["+ Add New..."])
            if b_cat == "+ Add New...":
                b_cat = st.text_input("New Category Name")
        with c2:
            b_val = st.number_input("Budget Amount (₹)", min_value=0.0, step=500.0)
            
            # --- UPDATED FREQUENCY SELECTOR (Now outside form for instant interaction) ---
            b_freq = st.radio("Budget Frequency", ["Entire Year (Apr-Mar)", "Specific Month"], horizontal=True)
            b_month = "April" # Default
            if b_freq == "Specific Month":
                # Auto-select current month if possible
                curr_m = calendar.month_name[datetime.date.today().month]
                def_idx = FY_MONTHS.index(curr_m) if curr_m in FY_MONTHS else 0
                b_month = st.selectbox("Select Month", FY_MONTHS, index=def_idx)
        
        if st.button("💾 Save Budget Goal"):
            if b_cat and b_val >= 0:
                try:
                    wb = openpyxl.load_workbook('Finance Tracker.xlsx')
                    # FORCE CLEAN STRUCTURE: Ensure sheet exists with correct headers
                    if 'Budget Planning' not in wb.sheetnames:
                        ws = wb.create_sheet('Budget Planning')
                        ws.append(['Type', 'Category'] + FY_MONTHS)
                    else: 
                        ws = wb['Budget Planning']
                        # Check if headers match standard
                        if ws.cell(row=1, column=3).value != "April": 
                            # If old structure, clear and reset header
                            ws.delete_rows(1, ws.max_row)
                            ws.append(['Type', 'Category'] + FY_MONTHS)
                    
                    target_row = None
                    for row in ws.iter_rows(min_row=2):
                        if row[0].value == b_type and row[1].value == b_cat:
                            target_row = row
                            break
                    
                    if not target_row:
                        ws.append([b_type, b_cat] + [0]*12)
                        target_row = tuple(ws.rows)[-1]
                    
                    if b_freq == "Entire Year (Apr-Mar)":
                        # Update all months (indices 2 to 13 correspond to Cols 3 to 14)
                        for i in range(2, 14): target_row[i].value = b_val
                    else:
                        # Map selected month to column index
                        # Header is: Type(0), Cat(1), April(2)...
                        month_idx_in_tuple = FY_MONTHS.index(b_month) + 2
                        target_row[month_idx_in_tuple].value = b_val
                    
                    wb.save('Finance Tracker.xlsx')
                    st.success(f"✅ Budget for {b_cat} updated!")
                    st.cache_data.clear(); st.rerun()
                except Exception as e: st.error(f"Error: {e}. Is Excel file open?")

    with tab_month_view:
        st.write("### Filter by Month")
        view_month = st.selectbox("Select Month to View", FY_MONTHS)
        
        # Filter the melted budget dataframe
        if not budget_df.empty:
            monthly_data = budget_df[budget_df['Month'] == view_month].copy()
            monthly_data = monthly_data[monthly_data['Amount'] > 0] # Show only active budgets
            
            if not monthly_data.empty:
                st.dataframe(monthly_data[['Type', 'Category', 'Amount']].style.format({'Amount': '₹{:,.0f}'}), use_container_width=True, hide_index=True)
            else:
                st.info(f"No budget set for {view_month}.")
        else:
            st.info("Budget data not loaded.")

    with tab_full_matrix:
        st.write("### Full Annual Matrix")
        # Just show the simple raw matrix from Excel
        if not budget_matrix_df.empty:
            st.dataframe(budget_matrix_df, use_container_width=True, hide_index=True)
        else:
            st.info("Budget matrix is empty.")

elif page == "📈 Investment Tracker":
    st.markdown("<h2>📈 Investment Tracker</h2>", unsafe_allow_html=True)
    summary_df = pd.DataFrame(); fd_calc = pd.DataFrame(); treemap_data = []; all_inv_timeline = []

    if 'Details' in df.columns:
        t_col = df['Details'].astype(str).str.extract(r'\[Ticker:\s*([^,\]]+)')[0]
        q_col = df['Details'].astype(str).str.extract(r'Qty:\s*([0-9.-]+)')[0]
        c_col = df['Details'].astype(str).str.extract(r'Class:\s*([^,\]]+)')[0]
        extracted = pd.DataFrame({'Ticker': t_col, 'Qty': q_col, 'Asset_Class': c_col})
        portfolio_data = pd.concat([df[['Date', 'Category', 'Amount']], extracted], axis=1).dropna(subset=['Ticker'])
        portfolio_data['Qty'] = pd.to_numeric(portfolio_data['Qty'], errors='coerce').fillna(0)
        portfolio_data.loc[portfolio_data['Asset_Class'].isna() & portfolio_data['Category'].isin(['Mutual Funds', 'Stocks / ETFs', 'Stocks', 'Stock', 'Equity']), 'Asset_Class'] = 'Equity'
        portfolio_data.loc[portfolio_data['Asset_Class'].isna() & portfolio_data['Category'].isin(['Crypto', 'Cryptocurrency']), 'Asset_Class'] = 'Crypto'
        portfolio_data['Asset_Class'] = portfolio_data['Asset_Class'].fillna('Other')
        if not portfolio_data.empty:
            all_inv_timeline.append(portfolio_data[['Date', 'Amount']]) 
            summary_df = portfolio_data.groupby('Ticker').agg(Total_Qty=('Qty', 'sum'), Total_Invested=('Amount', 'sum'), Asset_Class=('Asset_Class', 'first')).reset_index()
            summary_df = summary_df[summary_df['Total_Qty'] > 0]
            live_prices_dict, live_names_dict = get_market_data(summary_df['Ticker'].unique(), amfi_data_dict, nse_data_dict)
            summary_df['Asset Name'] = summary_df['Ticker'].map(live_names_dict)
            summary_df['Live Price (₹)'] = summary_df['Ticker'].map(live_prices_dict).fillna(0.0)
            summary_df['Current Value (₹)'] = summary_df['Live Price (₹)'] * summary_df['Total_Qty']
            summary_df['Unrealized P&L (₹)'] = summary_df['Current Value (₹)'] - summary_df['Total_Invested']
            for _, row in summary_df.iterrows():
                ret_pct = (row['Unrealized P&L (₹)'] / row['Total_Invested'] * 100) if row['Total_Invested'] > 0 else 0.0
                treemap_data.append({'Asset_Class': row['Asset_Class'], 'Asset_Name': row['Asset Name'], 'Value': row['Current Value (₹)'], 'Return_Pct': ret_pct})

        fd_df = df[df['Category'] == 'Fixed Deposit'].copy()
        if not fd_df.empty:
            fd_extracted = fd_df['Details'].astype(str).str.extract(r'ROI:\s*([0-9.]+)%,\s*Period:\s*([0-9]+)M')
            fd_extracted.columns = ['ROI', 'Period']
            fd_calc = pd.concat([fd_df[['Date', 'Amount']], fd_extracted], axis=1).dropna(subset=['ROI', 'Period'])
            if not fd_calc.empty:
                all_inv_timeline.append(fd_calc[['Date', 'Amount']]) 
                fd_calc['ROI'] = pd.to_numeric(fd_calc['ROI'])
                fd_calc['Period'] = pd.to_numeric(fd_calc['Period'])
                today = pd.Timestamp(datetime.datetime.today().date())
                fd_calc['Maturity Date'] = fd_calc.apply(lambda row: row['Date'] + pd.DateOffset(months=int(row['Period'])), axis=1)
                fd_calc['End Date'] = fd_calc['Maturity Date'].apply(lambda x: min(x, today)) 
                fd_calc['Elapsed Days'] = (fd_calc['End Date'] - fd_calc['Date']).dt.days
                fd_calc['Accrued Interest (₹)'] = fd_calc['Amount'] * (fd_calc['ROI'] / 100) * (fd_calc['Elapsed Days'] / 365)
                fd_calc['Current Value (₹)'] = fd_calc['Amount'] + fd_calc['Accrued Interest (₹)']
                for _, row in fd_calc.iterrows():
                    ret_pct = (row['Accrued Interest (₹)'] / row['Amount'] * 100) if row['Amount'] > 0 else 0.0
                    treemap_data.append({'Asset_Class': 'Debt', 'Asset_Name': f"FD: {row['ROI']}% {row['Period']}M", 'Value': row['Current Value (₹)'], 'Return_Pct': ret_pct})

    treemap_df = pd.DataFrame(treemap_data)
    if not treemap_df.empty:
        treemap_df = treemap_df[treemap_df['Value'] > 0]
        treemap_df['Return_Pct'] = treemap_df['Return_Pct'].fillna(0.0)

    grand_total_invested = sum([x['Value'] / (1 + x['Return_Pct']/100) for x in treemap_data]) if treemap_data else 0
    grand_total_current = treemap_df['Value'].sum() if not treemap_df.empty else 0

    portfolio_xirr = 0.0; nifty_benchmark_active = False; current_benchmark_val = 0.0
    if all_inv_timeline and grand_total_current > 0:
        all_inv_df = pd.concat(all_inv_timeline).sort_values('Date')
        cash_flows = [(row['Date'].date(), -row['Amount']) for _, row in all_inv_df.iterrows()]
        cash_flows.append((datetime.datetime.today().date(), grand_total_current))
        portfolio_xirr = calc_xirr(cash_flows) * 100
        earliest_date = all_inv_df['Date'].min()
        nifty_series = fetch_benchmark_history(earliest_date)
        if not nifty_series.empty:
            nifty_benchmark_active = True
            all_inv_df['Nifty_Price'] = all_inv_df['Date'].dt.normalize().map(nifty_series).ffill().bfill().fillna(10000) 
            all_inv_df['Nifty_Units'] = all_inv_df['Amount'] / all_inv_df['Nifty_Price']

    brok_inc = df[(df['Bank'] == 'Brokerage Account') & (df['Type'] == 'Income')]['Amount'].sum()
    brok_exp = df[(df['Bank'] == 'Brokerage Account') & (df['Type'] == 'Expenses')]['Amount'].sum()
    brok_sav = df[(df['Bank'] == 'Brokerage Account') & (df['Type'] == 'Savings')]['Amount'].sum()
    brokerage_cash = brok_inc - brok_exp - brok_sav

    st.markdown("### 📊 Performance Analytics")
    col_k1, col_k2, col_k3, col_k4 = st.columns(4)
    with col_k1: st.markdown(f"<div class='glow-card glow-blue'><div class='kpi-title'>PORTFOLIO VALUE</div><div class='kpi-value' style='color:#2962ff'>₹{grand_total_current:,.0f}</div></div>", unsafe_allow_html=True)
    with col_k2: 
        xirr_color = '#00c853' if portfolio_xirr >= 0 else '#d50000'
        xirr_label = "Overall XIRR"
        st.markdown(f"<div class='glow-card glow-green'><div class='kpi-title'>{xirr_label}</div><div class='kpi-value' style='color:{xirr_color}'>{portfolio_xirr:,.2f}%</div></div>", unsafe_allow_html=True)
    with col_k3:
        current_benchmark_val = all_inv_df['Nifty_Units'].sum() * nifty_series.iloc[-1] if nifty_benchmark_active else 0
        st.markdown(f"<div class='glow-card glow-gold'><div class='kpi-title'>NIFTY 50 BENCHMARK</div><div class='kpi-value' style='color:#ffd600'>₹{current_benchmark_val:,.0f}</div></div>", unsafe_allow_html=True)
    with col_k4:
        st.markdown(f"<div class='glow-card glow-blue'><div class='kpi-title'>BROKERAGE CASH</div><div class='kpi-value'>₹{brokerage_cash:,.0f}</div></div>", unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)
    st.markdown("<div class='chart-box'>", unsafe_allow_html=True)
    st.markdown("<h4 style='text-align: center;'>🌊 Wealth Flow Diagram</h4>", unsafe_allow_html=True)
    
    if not treemap_df.empty:
        class_group = treemap_df.groupby('Asset_Class')['Value'].sum().reset_index()
        all_classes = class_group['Asset_Class'].tolist(); all_assets = treemap_df['Asset_Name'].tolist()
        labels = ["Total Portfolio"] + all_classes + all_assets
        label_to_index = {label: i for i, label in enumerate(labels)}; display_labels = [str(l) for l in labels]
        base_colors = ['#ff006e', '#3a86ff', '#fb5607', '#8338ec', '#ffbe0b', '#00b4d8']
        class_color_map = {cls: base_colors[i % len(base_colors)] for i, cls in enumerate(all_classes)}
        node_colors = [root_node_color] 
        for cls in all_classes: node_colors.append(class_color_map[cls])
        for asset in all_assets:
            parent_cls = treemap_df[treemap_df['Asset_Name'] == asset]['Asset_Class'].iloc[0]
            node_colors.append(class_color_map[parent_cls])
        source, target, values, link_colors = [], [], [], []
        for _, row in class_group.iterrows():
            source.append(label_to_index["Total Portfolio"]); target.append(label_to_index[row['Asset_Class']]); values.append(row['Value'])
            hc = class_color_map[row['Asset_Class']].lstrip('#')
            link_colors.append(f"rgba({int(hc[0:2], 16)}, {int(hc[2:4], 16)}, {int(hc[4:6], 16)}, 0.4)")
        for _, row in treemap_df.iterrows():
            source.append(label_to_index[row['Asset_Class']]); target.append(label_to_index[row['Asset_Name']]); values.append(row['Value'])
            hc = class_color_map[row['Asset_Class']].lstrip('#')
            link_colors.append(f"rgba({int(hc[0:2], 16)}, {int(hc[2:4], 16)}, {int(hc[4:6], 16)}, 0.25)")
        fig_sankey = go.Figure(data=[go.Sankey(node = dict(pad=30, thickness=40, line=dict(color="gray", width=1), label=display_labels, color=node_colors), link = dict(source=source, target=target, value=values, color=link_colors))])
        fig_sankey.update_layout(height=550, margin=dict(t=30, b=30, l=30, r=30), paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)', font=dict(size=14, family="Arial, sans-serif", color=chart_text_color))
        st.plotly_chart(fig_sankey, use_container_width=True)
    else: st.info("No active assets to allocate.")
    st.markdown("</div>", unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)
    st.subheader("📡 Live Investment Screener")
    if not summary_df.empty:
        display_df = summary_df[['Asset Name', 'Total_Qty', 'Total_Invested', 'Live Price (₹)', 'Current Value (₹)', 'Unrealized P&L (₹)']]
        display_df = display_df.rename(columns={'Total_Qty': 'Total Units', 'Total_Invested': 'Invested Value (₹)'})
        def color_pnl(val):
            if pd.isna(val): return ''
            return 'color: #00c853; font-weight: bold;' if val > 0 else 'color: #d50000; font-weight: bold;' if val < 0 else 'color: gray; font-weight: bold;'
        styled_screener = display_df.style.map(color_pnl, subset=['Unrealized P&L (₹)']).format({'Total Units': '{:.4f}', 'Invested Value (₹)': '₹{:,.2f}', 'Live Price (₹)': '₹{:,.2f}', 'Current Value (₹)': '₹{:,.2f}', 'Unrealized P&L (₹)': '₹{:,.2f}'})
        st.dataframe(styled_screener, use_container_width=True, hide_index=True)
    else: st.info("No active stocks, mutual funds, or crypto tracked yet.")
    
    st.markdown("<br>", unsafe_allow_html=True)
    st.markdown("#### 🏦 Active Fixed Deposits")
    if not fd_calc.empty:
        fd_display = fd_calc[['Date', 'Amount', 'ROI', 'Period', 'Maturity Date', 'Accrued Interest (₹)', 'Current Value (₹)']].copy()
        fd_display.rename(columns={'Date': 'Start Date', 'Amount': 'Invested Value (₹)', 'Period': 'Tenure (Months)', 'ROI': 'ROI (%)'}, inplace=True)
        today = pd.Timestamp(datetime.datetime.today().date())
        def highlight_matured(val): return 'color: #00c853; font-weight: bold;' if val <= today else ''
        styled_fd = fd_display.style.map(highlight_matured, subset=['Maturity Date']).format({'Start Date': lambda x: x.strftime('%Y-%m-%d'), 'Maturity Date': lambda x: x.strftime('%Y-%m-%d'), 'Invested Value (₹)': '₹{:,.2f}', 'ROI (%)': '{:.2f}%', 'Accrued Interest (₹)': '₹{:,.2f}', 'Current Value (₹)': '₹{:,.2f}'})
        st.dataframe(styled_fd, use_container_width=True, hide_index=True)
        
    st.markdown("---")
    st.subheader("💸 Sell Investment & Tax Calculator")
    if not summary_df.empty:
        with st.expander("📉 Sell an Asset & Route Cash to Brokerage Account"):
            sell_ticker = st.selectbox("Select Asset to Sell", summary_df['Ticker'].tolist())
            asset_info = summary_df[summary_df['Ticker'] == sell_ticker].iloc[0]
            asset_history = portfolio_data[(portfolio_data['Ticker'] == sell_ticker) & (portfolio_data['Amount'] > 0)]
            earliest_date = asset_history['Date'].min() if not asset_history.empty else datetime.datetime.today()
            avg_buy_price = asset_info['Total_Invested'] / asset_info['Total_Qty'] if asset_info['Total_Qty'] > 0 else 0
            st.info(f"**{asset_info['Asset Name']}** | Available Qty: {asset_info['Total_Qty']:.4f} | Avg Buy Price: ₹{avg_buy_price:,.2f} | First Bought: {earliest_date.strftime('%Y-%m-%d')}")
            sell_col1, sell_col2, sell_col3 = st.columns(3)
            with sell_col1: sell_qty = st.number_input("Quantity to Sell", min_value=0.0001, max_value=float(asset_info['Total_Qty']), step=1.0)
            with sell_col2: sell_rate = st.number_input("Sell Rate / NAV (₹)", min_value=0.01, value=float(asset_info['Live Price (₹)']))
            with sell_col3: sell_date = st.date_input("Date of Sale", datetime.datetime.today())
            gross_sale_value = sell_qty * sell_rate
            buy_cost = sell_qty * avg_buy_price
            profit = gross_sale_value - buy_cost
            holding_months = (pd.to_datetime(sell_date).date() - earliest_date.date()).days / 30.41
            tax_amt = 0; tax_type = "Loss (No Tax)"; asset_class = asset_info['Asset_Class']
            if profit > 0:
                if asset_class == 'Crypto': tax_amt = profit * 0.30; tax_type = "Crypto Flat Tax (30%)"
                elif asset_class == 'Equity':
                    if holding_months > 12: tax_amt = profit * 0.125; tax_type = "LTCG @ 12.5% (>12 Months)"
                    else: tax_amt = profit * 0.20; tax_type = "STCG @ 20% (<=12 Months)"
                else: 
                    if holding_months > 24: tax_amt = profit * 0.125; tax_type = "LTCG @ 12.5% (>24 Months)"
                    else: tax_amt = profit * 0.30; tax_type = "STCG @ Slab Rate (~30%)"
            net_proceeds = gross_sale_value - tax_amt
            st.markdown(f"<div class='glow-card'><h4 style='margin-bottom:5px; color:gray;'>NET PROCEEDS</h4><h2 style='color:#2962ff;'>₹{net_proceeds:,.2f}</h2><span style='color:{'#00c853' if profit >= 0 else '#d50000'}'>Profit: ₹{profit:,.2f}</span> | <span style='color:#d50000'>Tax: ₹{tax_amt:,.2f}</span></div>", unsafe_allow_html=True)
            if st.button("🚀 Execute Sale & Save to Ledger"):
                try:
                    wb = openpyxl.load_workbook('Finance Tracker.xlsx'); sheet = wb['Budget Tracking']
                    next_row = 11
                    while sheet.cell(row=next_row, column=2).value: next_row += 1
                    orig_category = asset_history['Category'].mode()[0] if not asset_history.empty else "Stocks / ETFs"
                    sheet.cell(row=next_row, column=2, value=sell_date.strftime("%Y-%m-%d")); sheet.cell(row=next_row, column=3, value="Savings"); sheet.cell(row=next_row, column=4, value=orig_category); sheet.cell(row=next_row, column=5, value=-buy_cost); sheet.cell(row=next_row, column=6, value=f"[Bank: Brokerage Account] [Ticker: {sell_ticker}, Qty: {-sell_qty}, Class: {asset_class}] (Principal Recovered)")
                    next_row += 1
                    sheet.cell(row=next_row, column=2, value=sell_date.strftime("%Y-%m-%d"))
                    if profit >= 0: sheet.cell(row=next_row, column=3, value="Income"); sheet.cell(row=next_row, column=4, value="Investment Payout"); sheet.cell(row=next_row, column=5, value=(profit - tax_amt))
                    else: sheet.cell(row=next_row, column=3, value="Expenses"); sheet.cell(row=next_row, column=4, value="Investment Loss"); sheet.cell(row=next_row, column=5, value=(abs(profit) + tax_amt))
                    sheet.cell(row=next_row, column=6, value=f"[Bank: Brokerage Account] Profit/Loss from {sell_qty} {sell_ticker}. Tax Deducted: {tax_amt:.0f} ({tax_type})")
                    wb.save('Finance Tracker.xlsx'); st.success(f"✅ Sold {sell_ticker}! Money is now in your Brokerage Account."); get_market_data.clear(); st.rerun()
                except Exception as e: st.error(f"Error: {e}. Is Excel open?")


elif page == "💳 Credit Cards":
    st.markdown("<h2>💳 Credit Health Command Center</h2>", unsafe_allow_html=True)
    st.write("Monitor your live credit utilization, track upcoming due dates, and analyze your spending DNA across all your cards.")
    ALL_AVAILABLE_CARDS = sorted(list(set(["Airtel Axis Bank Credit Card", "Axis Atlas Credit Card", "HDFC Infinia Credit Card", "HDFC Diners Club Black", "HDFC Regalia Gold Credit Card", "Amazon Pay ICICI Bank Credit Card", "SBI Cashback Card", "OneCard", "ICICI Bank HPCL Super Saver Credit Card", "IndianOil HDFC Bank Credit Card"]))) 
    with st.expander("⚙️ Manage Wallet & Goals (Add / Edit Cards)"):
        st.write("Set your Statement Date (1-31) to track due dates, and your Annual Target Spend to track fee waivers.")
        if cc_df.empty:
            cc_edit_df = pd.DataFrame(columns=['Card Name', 'Total Limit', 'Statement Date', 'Target Spend (₹)'])
            combined_options = ALL_AVAILABLE_CARDS
        else:
            cc_edit_df = cc_df.copy()
            for col in ['Statement Date', 'Target Spend (₹)']:
                if col not in cc_edit_df.columns:
                    cc_edit_df[col] = 15 if col == 'Statement Date' else 100000
            cc_edit_df['Card Name'] = cc_edit_df['Card Name'].astype(str).str.strip()
            cc_edit_df['Card Name'] = cc_edit_df['Card Name'].replace('nan', '')
            existing_names = cc_edit_df['Card Name'].tolist()
            combined_options = sorted(list(set(ALL_AVAILABLE_CARDS + existing_names)))
            if "" in combined_options: combined_options.remove("")
        edited_cc = st.data_editor(cc_edit_df, num_rows="dynamic", use_container_width=True,
            column_config={"Card Name": st.column_config.SelectboxColumn("Card Name", options=combined_options), "Total Limit": st.column_config.NumberColumn("Total Limit (₹)", min_value=0, step=10000), "Statement Date": st.column_config.NumberColumn("Statement Day (1-31)", min_value=1, max_value=31, step=1), "Target Spend (₹)": st.column_config.NumberColumn("Annual Target Spend (₹)", min_value=0, step=50000)})
        if st.button("💾 Save Wallet Data"):
            try:
                wb = openpyxl.load_workbook('Finance Tracker.xlsx')
                if 'Credit Cards' in wb.sheetnames: del wb['Credit Cards']
                ws = wb.create_sheet('Credit Cards') 
                ws.append(['Card Name', 'Total Limit', 'Statement Date', 'Target Spend (₹)'])
                for _, row in edited_cc.iterrows():
                    if pd.notna(row['Card Name']) and str(row['Card Name']).strip() != "":
                        ws.append([str(row['Card Name']).strip(), row['Total Limit'], row['Statement Date'], row['Target Spend (₹)']])
                wb.save('Finance Tracker.xlsx'); st.success("✅ Wallet settings securely saved!"); st.rerun()
            except Exception as e: st.error(f"Error saving: {e}. Please close your Excel file.")

    owned_cards = cc_df['Card Name'].dropna().unique().tolist() if not cc_df.empty else []
    if not owned_cards: st.info("👆 Please add your credit cards in the Wallet Manager above to unlock your analytics.")
    else:
        st.markdown("---")
        selected_card = st.selectbox("🔍 Select a Card to Analyze", owned_cards)
        card_info = cc_df[cc_df['Card Name'] == selected_card].iloc[0]
        c_limit = float(card_info['Total Limit'])
        c_stmt_day = int(card_info.get('Statement Date', 15))
        c_target = float(card_info.get('Target Spend (₹)', 100000))
        card_txns = df[df['Bank'] == selected_card].copy()
        spent = card_txns[card_txns['Type'] == 'Expenses']['Amount'].sum()
        paid = card_txns[card_txns['Type'] == 'Income']['Amount'].sum()
        outstanding = max(0, spent - paid)
        available = c_limit - outstanding
        utilization_pct = (outstanding / c_limit * 100) if c_limit > 0 else 0
        util_color = "#00c853" 
        if utilization_pct > 30: util_color = "#ffd600" 
        if utilization_pct > 50: util_color = "#ff6d00" 
        if utilization_pct > 75: util_color = "#d50000" 
        today = datetime.date.today()
        if today.day >= c_stmt_day: last_stmt = datetime.date(today.year, today.month, c_stmt_day)
        else:
            prev_month = today.month - 1 if today.month > 1 else 12
            prev_year = today.year if today.month > 1 else today.year - 1
            try: last_stmt = datetime.date(prev_year, prev_month, c_stmt_day)
            except ValueError: last_stmt = datetime.date(prev_year, prev_month, 28) 
        due_date = last_stmt + datetime.timedelta(days=20)
        days_to_due = (due_date - today).days
        due_str = f"Due in {days_to_due} Days" if days_to_due >= 0 else f"Overdue by {abs(days_to_due)} Days!"
        due_color = "#d50000" if days_to_due < 5 else "#2962ff"
        current_year = today.year
        ytd_spend = card_txns[(card_txns['Type'] == 'Expenses') & (card_txns['Year'] == current_year)]['Amount'].sum()
        milestone_pct = min((ytd_spend / c_target * 100), 100) if c_target > 0 else 0

        st.markdown(f"<h3 style='text-align:center; color:#e0e0e0;'>{selected_card}</h3>", unsafe_allow_html=True)
        k1, k2, k3, k4 = st.columns(4)
        with k1: st.markdown(f"<div class='glow-card'><div class='kpi-title'>Total Outstanding</div><div class='kpi-value' style='color:#ff1744'>₹{outstanding:,.0f}</div></div>", unsafe_allow_html=True)
        with k2: st.markdown(f"<div class='glow-card'><div class='kpi-title'>Available Limit</div><div class='kpi-value' style='color:#00c853'>₹{available:,.0f}</div></div>", unsafe_allow_html=True)
        with k3: st.markdown(f"<div class='glow-card'><div class='kpi-title'>Statement Day</div><div class='kpi-value' style='color:gray'>{c_stmt_day}th</div></div>", unsafe_allow_html=True)
        with k4: st.markdown(f"<div class='glow-card'><div class='kpi-title'>Est. Payment Date</div><div class='kpi-value' style='color:{due_color}'>{due_str}</div></div>", unsafe_allow_html=True)

        c_chart1, c_chart2 = st.columns(2)
        with c_chart1:
            st.markdown("<div class='chart-box'><h5 style='text-align: center; color:gray;'>CIBIL UTILIZATION MONITOR</h5>", unsafe_allow_html=True)
            fig_gauge = go.Figure(go.Indicator(mode = "gauge+number", value = utilization_pct, number = {'suffix': "%", 'font': {'color': util_color}}, domain = {'x': [0, 1], 'y': [0, 1]}, gauge = {'axis': {'range': [None, 100], 'tickwidth': 1, 'tickcolor': "gray"}, 'bar': {'color': util_color}, 'bgcolor': "rgba(128,128,128,0.1)", 'borderwidth': 2, 'bordercolor': "gray", 'steps': [{'range': [0, 30], 'color': "rgba(0, 200, 83, 0.1)"}, {'range': [30, 50], 'color': "rgba(255, 214, 0, 0.1)"}, {'range': [50, 100], 'color': "rgba(213, 0, 0, 0.1)"}], 'threshold': {'line': {'color': "red", 'width': 4}, 'thickness': 0.75, 'value': 30}}))
            fig_gauge.update_layout(height=300, margin=dict(t=30, b=10, l=10, r=10), paper_bgcolor='rgba(0,0,0,0)', font=dict(color=chart_text_color))
            st.plotly_chart(fig_gauge, use_container_width=True)
            if utilization_pct > 30: st.warning("⚠️ High Utilization! Spending over 30% of your limit can lower your CIBIL score.")
            st.markdown("</div>", unsafe_allow_html=True)

        with c_chart2:
            st.markdown("<div class='chart-box'><h5 style='text-align: center; color:gray;'>CARD SPEND DNA</h5>", unsafe_allow_html=True)
            card_exp = card_txns[(card_txns['Type'] == 'Expenses') & (card_txns['Amount'] > 0)]
            if not card_exp.empty:
                dna_df = process_pie_data(card_exp, 0.05)
                fig_dna = px.pie(dna_df, values='Amount', names='Category', hole=0.7)
                fig_dna.update_layout(height=300, paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)', font=dict(color=chart_text_color), margin=dict(t=10, b=10, l=10, r=10))
                fig_dna.update_traces(textposition='inside', textinfo='percent+label', marker=dict(colors=px.colors.sequential.Agsunset))
                st.plotly_chart(fig_dna, use_container_width=True)
            else: st.info("No expenses tracked on this card yet.")
            st.markdown("</div>", unsafe_allow_html=True)

        st.markdown("<div class='chart-box'><h5 style='color:gray;'>🎯 ANNUAL MILESTONE & FEE WAIVER TRACKER</h5>", unsafe_allow_html=True)
        st.write(f"**Year-to-Date Spend:** ₹{ytd_spend:,.0f} / **Target:** ₹{c_target:,.0f}")
        st.progress(int(milestone_pct))
        if milestone_pct >= 100: st.success("🎉 Goal Reached! You have cleared the target spend for this card.")
        else: st.info(f"Spend ₹{c_target - ytd_spend:,.0f} more this year to hit your milestone.")
        st.markdown("</div>", unsafe_allow_html=True)


elif page == "🔄 Subscription Radar":
    st.markdown("<h2>🔄 Subscription Radar</h2>", unsafe_allow_html=True)
    st.write("Track all your fixed auto-debits. (Hint: Subscriptions are automatically added here when you check 'Recurring Subscription' while adding a new expense).")
    
    with st.expander("⚙️ Manage Master Subscriptions list"):
        if subs_df.empty:
            subs_edit_df = pd.DataFrame(columns=['Service Name', 'Category', 'Billing Cycle', 'Amount (₹)', 'Next Due Date'])
        else:
            subs_edit_df = subs_df.copy()
            subs_edit_df['Next Due Date'] = pd.to_datetime(subs_edit_df['Next Due Date']).dt.date
            
        edited_subs = st.data_editor(
            subs_edit_df,
            num_rows="dynamic",
            use_container_width=True,
            column_config={
                "Category": st.column_config.SelectboxColumn("Category", options=["Entertainment", "Software", "Gym/Fitness", "Utilities", "Other"]),
                "Billing Cycle": st.column_config.SelectboxColumn("Billing Cycle", options=["Monthly", "Yearly", "Quarterly"]),
                "Next Due Date": st.column_config.DateColumn("Next Due Date")
            }
        )
        
        if st.button("💾 Save Subscriptions"):
            try:
                wb = openpyxl.load_workbook('Finance Tracker.xlsx')
                if 'Subscriptions' in wb.sheetnames: del wb['Subscriptions']
                ws = wb.create_sheet('Subscriptions')
                ws.append(['Service Name', 'Category', 'Billing Cycle', 'Amount (₹)', 'Next Due Date'])
                for _, row in edited_subs.iterrows():
                    if pd.notna(row['Service Name']):
                        ws.append([row['Service Name'], row['Category'], row['Billing Cycle'], row['Amount (₹)'], str(row['Next Due Date']) if pd.notna(row['Next Due Date']) else ""])
                wb.save('Finance Tracker.xlsx')
                st.success("✅ Subscriptions updated!")
                st.rerun()
            except Exception as e:
                st.error(f"Error saving: {e}. Close Excel.")

    if not subs_df.empty:
        monthly_subs = subs_df[subs_df['Billing Cycle'] == 'Monthly']['Amount (₹)'].sum()
        yearly_subs = subs_df[subs_df['Billing Cycle'] == 'Yearly']['Amount (₹)'].sum()
        avg_monthly = monthly_subs + (yearly_subs / 12)
        
        col_s1, col_s2, col_s3 = st.columns(3)
        with col_s1: st.markdown(f"<div class='glow-card glow-purple'><div class='kpi-title'>AVG MONTHLY BURN</div><div class='kpi-value' style='color:#b388ff'>₹{avg_monthly:,.0f}</div></div>", unsafe_allow_html=True)
        with col_s2: st.markdown(f"<div class='glow-card'><div class='kpi-title'>TOTAL ACTIVE SUBS</div><div class='kpi-value'>{len(subs_df)}</div></div>", unsafe_allow_html=True)
        with col_s3: st.markdown(f"<div class='glow-card'><div class='kpi-title'>YEARLY COMMITMENT</div><div class='kpi-value' style='color:gray'>₹{avg_monthly*12:,.0f}</div></div>", unsafe_allow_html=True)
        
        st.markdown("### Upcoming Auto-Debits")
        subs_display = subs_df.copy()
        subs_display['Next Due Date'] = pd.to_datetime(subs_display['Next Due Date'])
        subs_display = subs_display.sort_values('Next Due Date')
        
        today = pd.Timestamp(datetime.date.today())
        def highlight_due(row):
            if pd.isna(row['Next Due Date']): return [''] * len(row)
            days_left = (row['Next Due Date'] - today).days
            if days_left <= 7: return ['background-color: rgba(255, 23, 68, 0.2); color: #ff1744'] * len(row)
            return [''] * len(row)
            
        subs_display['Next Due Date'] = subs_display['Next Due Date'].dt.strftime('%d %b %Y')
        st.dataframe(subs_display.style.apply(highlight_due, axis=1).format({'Amount (₹)': '₹{:,.0f}'}), use_container_width=True, hide_index=True)
    else:
        st.info("No subscriptions tracked yet. Add them via the sidebar or in the manager above.")

elif page == "🤝 Splitwise / Settles":
    st.markdown("<h2>🤝 Household Split Engine</h2>", unsafe_allow_html=True)
    st.write("Manage shared expenses. (Hint: You can instantly create a split when adding an expense in the left sidebar!).")
    col_sp1, col_sp2 = st.columns([1, 2])
    with col_sp1:
        st.markdown("<div class='chart-box'>", unsafe_allow_html=True)
        st.markdown("#### ➕ Add Shared Expense")
        sp_date = st.date_input("Date", datetime.datetime.today())
        sp_desc = st.text_input("What was it? (e.g., Dinner, Groceries)")
        sp_total = st.number_input("Total Bill Amount (₹)", min_value=0.0)
        sp_other_person = st.selectbox("Who are you splitting with?", split_users + ["+ Add New Person..."])
        if sp_other_person == "+ Add New Person...": sp_other_person = st.text_input("Enter Person's Name")
        sp_payer = st.selectbox("Who Paid?", ["Jaynik", sp_other_person] if sp_other_person else ["Jaynik"])
        sp_debtor = sp_other_person if sp_payer == "Jaynik" else "Jaynik"
        sp_split = st.number_input(f"Amount {sp_debtor} Owes (₹)", min_value=0.0, value=sp_total/2 if sp_total > 0 else 0.0)
        
        if st.button("💾 Log Split Expense", use_container_width=True):
            if sp_desc and sp_total > 0 and sp_other_person:
                try:
                    wb = openpyxl.load_workbook('Finance Tracker.xlsx')
                    if 'Splitwise' not in wb.sheetnames:
                        ws = wb.create_sheet('Splitwise')
                        ws.append(['ID', 'Date', 'Payer', 'Debtor', 'Total Amount', 'Split Amount', 'Description', 'Status'])
                    else: ws = wb['Splitwise']
                    split_id = f"SPL-{datetime.datetime.now().strftime('%Y%m%d%H%M%S')}"
                    ws.append([split_id, sp_date.strftime("%Y-%m-%d"), sp_payer, sp_debtor, sp_total, sp_split, sp_desc, "Pending"])
                    wb.save('Finance Tracker.xlsx')
                    st.success("✅ Split expense recorded!"); st.rerun()
                except Exception as e: st.error(f"Error: {e}")
        st.markdown("</div>", unsafe_allow_html=True)
        
    with col_sp2:
        st.markdown(f"### Live Balances for {current_user}")
        if not split_df.empty:
            pending_df = split_df[split_df['Status'] == 'Pending']
            owed_to_me = pending_df[pending_df['Payer'] == current_user]['Split Amount'].sum()
            i_owe = pending_df[pending_df['Debtor'] == current_user]['Split Amount'].sum()
            net_split_balance = owed_to_me - i_owe
            
            b1, b2, b3 = st.columns(3)
            with b1: st.markdown(f"<div class='glow-card'><div class='kpi-title'>OWED TO ME</div><div class='kpi-value' style='color:#00c853'>₹{owed_to_me:,.0f}</div></div>", unsafe_allow_html=True)
            with b2: st.markdown(f"<div class='glow-card'><div class='kpi-title'>I OWE</div><div class='kpi-value' style='color:#d50000'>₹{i_owe:,.0f}</div></div>", unsafe_allow_html=True)
            bal_color = "#00c853" if net_split_balance >= 0 else "#d50000"
            bal_text = "YOU ARE OWED" if net_split_balance >= 0 else "YOU OWE"
            with b3: st.markdown(f"<div class='glow-card'><div class='kpi-title'>{bal_text}</div><div class='kpi-value' style='color:{bal_color}'>₹{abs(net_split_balance):,.0f}</div></div>", unsafe_allow_html=True)
            
            st.markdown("#### Pending Settlements")
            display_pending = pending_df[['Date', 'Description', 'Payer', 'Debtor', 'Split Amount']].copy()
            st.dataframe(display_pending.style.format({'Split Amount': '₹{:,.2f}'}), use_container_width=True, hide_index=True)
            
            st.markdown("#### ✅ Settle Up")
            settle_id = st.selectbox("Select Transaction to Settle", pending_df['ID'].tolist(), format_func=lambda x: f"{pending_df[pending_df['ID']==x]['Description'].iloc[0]} (₹{pending_df[pending_df['ID']==x]['Split Amount'].iloc[0]})")
            if st.button("Mark as Settled"):
                try:
                    wb = openpyxl.load_workbook('Finance Tracker.xlsx')
                    sheet = wb['Splitwise']
                    for row in range(2, sheet.max_row + 1):
                        if sheet.cell(row=row, column=1).value == settle_id:
                            sheet.cell(row=row, column=8).value = "Settled"
                            break
                    wb.save('Finance Tracker.xlsx')
                    st.success("✅ Marked as settled!"); st.rerun()
                except Exception as e: st.error(f"Error: {e}")
        else:
            st.info("No shared expenses recorded yet.")


elif page == "⚖️ Net Worth & Goals":
    st.markdown("<h2>⚖️ Net Worth & Goals</h2>", unsafe_allow_html=True)
    
    total_cash = df[df['Type'] == 'Income']['Amount'].sum() - df[df['Type'] == 'Expenses']['Amount'].sum() - df[df['Type'] == 'Savings']['Amount'].sum()
    total_market_value = 0
    t_col = df['Details'].astype(str).str.extract(r'\[Ticker:\s*([^,\]]+)')[0]
    q_col = df['Details'].astype(str).str.extract(r'Qty:\s*([0-9.-]+)')[0]
    inv_df = pd.DataFrame({'Ticker': t_col, 'Qty': q_col})
    inv_df['Qty'] = pd.to_numeric(inv_df['Qty'], errors='coerce').fillna(0)
    inv_df = inv_df.dropna(subset=['Ticker'])
    if not inv_df.empty:
        inv_grouped = inv_df.groupby('Ticker')['Qty'].sum().reset_index()
        inv_grouped = inv_grouped[inv_grouped['Qty'] > 0]
        prices, _ = get_market_data(inv_grouped['Ticker'].unique(), amfi_data_dict, nse_data_dict)
        inv_grouped['Value'] = inv_grouped['Ticker'].map(prices) * inv_grouped['Qty']
        total_market_value = inv_grouped['Value'].sum()
    total_fd_value = 0
    fd_df = df[df['Category'] == 'Fixed Deposit'].copy()
    if not fd_df.empty:
        fd_extracted = fd_df['Details'].astype(str).str.extract(r'ROI:\s*([0-9.]+)%,\s*Period:\s*([0-9]+)M')
        fd_extracted.columns = ['ROI', 'Period']
        fd_calc = pd.concat([fd_df[['Date', 'Amount']], fd_extracted], axis=1).dropna(subset=['ROI', 'Period'])
        if not fd_calc.empty:
            fd_calc['ROI'] = pd.to_numeric(fd_calc['ROI'])
            fd_calc['Period'] = pd.to_numeric(fd_calc['Period'])
            today = pd.Timestamp(datetime.datetime.today().date())
            fd_calc['Maturity Date'] = fd_calc.apply(lambda row: row['Date'] + pd.DateOffset(months=int(row['Period'])), axis=1)
            fd_calc['End Date'] = fd_calc['Maturity Date'].apply(lambda x: min(x, today)) 
            fd_calc['Elapsed Days'] = (fd_calc['End Date'] - fd_calc['Date']).dt.days
            total_fd_value = (fd_calc['Amount'] + (fd_calc['Amount'] * (fd_calc['ROI'] / 100) * (fd_calc['Elapsed Days'] / 365))).sum()
    total_physical_assets = assets_df['Estimated Value (₹)'].sum() if not assets_df.empty else 0
    
    splitwise_asset = 0; splitwise_liability = 0
    if not split_df.empty and current_user != "Household (Combined)":
        pending = split_df[split_df['Status'] == 'Pending']
        splitwise_asset = pending[pending['Payer'] == current_user]['Split Amount'].sum()
        splitwise_liability = pending[pending['Debtor'] == current_user]['Split Amount'].sum()
    
    total_assets = total_cash + total_market_value + total_fd_value + total_physical_assets + splitwise_asset
    
    total_cc_debt = 0
    if not cc_df.empty:
        for card in cc_df['Card Name'].unique():
            spent = df[(df['Bank'] == card) & (df['Type'] == 'Expenses')]['Amount'].sum()
            paid = df[(df['Bank'] == card) & (df['Type'] == 'Income')]['Amount'].sum()
            outstanding = spent - paid
            if outstanding > 0: total_cc_debt += outstanding
    total_loan_debt = 0
    active_loans = []
    def calc_loan_schedule(principal, rate_pa, tenure_months, start_date, prepay_amt=0, prepay_month=1):
        if principal <= 0 or rate_pa <= 0 or tenure_months <= 0: return 0, 0, pd.DataFrame()
        r = (rate_pa / 100) / 12
        emi = principal * r * ((1 + r)**tenure_months) / (((1 + r)**tenure_months) - 1)
        today = datetime.datetime.today().date()
        start = pd.to_datetime(start_date).date()
        elapsed_months = (today.year - start.year) * 12 + (today.month - start.month)
        if elapsed_months < 0: elapsed_months = 0
        schedule = []
        rem_prin = principal
        for m in range(1, int(tenure_months) + 1):
            if rem_prin <= 0: break
            if m == prepay_month and prepay_amt > 0: rem_prin -= prepay_amt
            int_paid = rem_prin * r; prin_paid = emi - int_paid
            if prin_paid > rem_prin: prin_paid = rem_prin
            rem_prin -= prin_paid
            schedule.append({'Month': m, 'EMI': emi, 'Principal Paid': prin_paid, 'Interest Paid': int_paid, 'Remaining Principal': max(0, rem_prin)})
        sched_df = pd.DataFrame(schedule)
        current_out = principal if elapsed_months == 0 else (sched_df.iloc[elapsed_months-1]['Remaining Principal'] if elapsed_months <= len(sched_df) else 0)
        return emi, current_out, sched_df
    if not loan_df.empty:
        for _, row in loan_df.iterrows():
            emi, current_out, _ = calc_loan_schedule(row['Principal'], row['Interest Rate'], row['Tenure'], row['Start Date'])
            total_loan_debt += current_out
            active_loans.append({'Loan Name': row['Loan Name'], 'Type': row['Type'], 'EMI (₹)': emi, 'Outstanding Principal (₹)': current_out, 'Interest Rate': f"{row['Interest Rate']}%"})
            
    total_liabilities = total_cc_debt + total_loan_debt + splitwise_liability
    true_net_worth = total_assets - total_liabilities
    
    col_nw1, col_nw2, col_nw3 = st.columns(3)
    with col_nw1: st.markdown(f"<div class='glow-card glow-green'><div class='kpi-title'>TOTAL ASSETS</div><div class='kpi-value' style='color:#00c853'>₹{total_assets:,.0f}</div></div>", unsafe_allow_html=True)
    with col_nw2: st.markdown(f"<div class='glow-card glow-red'><div class='kpi-title'>TOTAL LIABILITIES</div><div class='kpi-value' style='color:#d50000'>₹{total_liabilities:,.0f}</div></div>", unsafe_allow_html=True)
    with col_nw3: st.markdown(f"<div class='glow-card glow-blue'><div class='kpi-title'>TRUE NET WORTH</div><div class='kpi-value' style='color:#2962ff'>₹{true_net_worth:,.0f}</div></div>", unsafe_allow_html=True)
    
    st.markdown("---")
    
    # --- 🎯 GOAL TRACKING SECTION (FIXED VALUE LOGIC) ---
    st.markdown("### 🎯 Goal Tracking")
    
    # OPTION 1: Add New Goal
    with st.expander("➕ Add New Goal", expanded=False):
        ng_col1, ng_col2 = st.columns(2)
        with ng_col1:
            add_goal_name = st.text_input("Goal Name")
            add_goal_target = st.number_input("Target Amount (₹)", min_value=0.0)
        with ng_col2:
            add_goal_date = st.date_input("Target Completion Date")
            add_goal_priority = st.slider("Priority", 1, 10, 5)
        
        if st.button("Save New Goal"):
            if add_goal_name and add_goal_target > 0:
                try:
                    wb = openpyxl.load_workbook('Finance Tracker.xlsx')
                    if 'Goals' not in wb.sheetnames:
                        ws_goals = wb.create_sheet('Goals')
                        ws_goals.append(['Goal Name', 'Target Amount', 'Target Date', 'Priority', 'Status'])
                    else: ws_goals = wb['Goals']
                    ws_goals.append([add_goal_name, add_goal_target, add_goal_date.strftime("%Y-%m-%d"), add_goal_priority, "Active"])
                    wb.save('Finance Tracker.xlsx')
                    st.cache_data.clear() 
                    st.success("✅ Goal Created! Reloading...")
                    st.rerun()
                except Exception as e: st.error(f"Error saving goal: {e}")

    # OPTION 2: View Existing Goals
    if not goals_df.empty:
        # Calculate Portfolio XIRR for projection
        all_inv_df = df[df['Type'] == 'Savings'].copy()
        if not all_inv_df.empty:
            cash_flows = [(row['Date'].date(), -row['Amount']) for _, row in all_inv_df.iterrows()]
            cash_flows.append((datetime.datetime.today().date(), total_market_value + total_fd_value)) 
            port_xirr = calc_xirr(cash_flows)
            if port_xirr == 0.0: port_xirr = 0.10 # Fallback 10%
        else: port_xirr = 0.10 
        
        for idx, goal in goals_df.iterrows():
            with st.expander(f"📌 {goal['Goal Name']} (Target: ₹{goal['Target Amount']:,.0f})", expanded=True):
                # Filter investments linked to this goal
                goal_txns = df[df['Linked Goal'] == goal['Goal Name']]
                invested_corpus = goal_txns['Amount'].sum()
                
                # Accrued Logic:
                # 1. Start with raw principal invested
                current_goal_val = invested_corpus
                
                # 2. Add FD Interest logic specifically for this goal's FDs
                goal_fds = goal_txns[goal_txns['Category'] == 'Fixed Deposit']
                if not goal_fds.empty:
                    # Re-calculate FD accrual for specific goal items
                    for _, fd_row in goal_fds.iterrows():
                        # Extract ROI and Period from details string
                        roi_match = re.search(r'ROI:\s*([0-9.]+)%', str(fd_row['Details']))
                        roi_val = float(roi_match.group(1)) if roi_match else 0.0
                        elapsed = (pd.Timestamp(datetime.datetime.today().date()) - fd_row['Date']).days
                        interest_gained = fd_row['Amount'] * (roi_val / 100) * (elapsed / 365)
                        current_goal_val += interest_gained # Add interest to value
                
                # Projection
                target_date = pd.to_datetime(goal['Target Date'])
                days_remaining = (target_date - datetime.datetime.today()).days
                projected_val = current_goal_val * ((1 + port_xirr) ** (days_remaining / 365)) if days_remaining > 0 else current_goal_val
                
                gap = goal['Target Amount'] - projected_val
                status_color = "#00c853" if gap <= 0 else "#ff1744"
                status_msg = "✅ On Track" if gap <= 0 else f"⚠️ Shortfall of ₹{gap:,.0f}"
                
                c1, c2, c3 = st.columns(3)
                c1.metric("Invested so far", f"₹{invested_corpus:,.0f}")
                c2.metric("Current Value (Est.)", f"₹{current_goal_val:,.0f}")
                c3.markdown(f"<h4 style='color:{status_color}'>{status_msg}</h4>", unsafe_allow_html=True)
                
                st.progress(min(int((current_goal_val / goal['Target Amount']) * 100), 100))
                st.caption(f"Projected Value by {target_date.date()}: ₹{projected_val:,.0f} (Assumed Return: {port_xirr*100:.1f}%)")
    else:
        st.info("No goals set yet. Add a goal above!")

    st.markdown("---")
    c_as, c_li = st.columns(2)
    with c_as:
        st.markdown("#### 🟢 Asset Breakdown")
        st.write(f"- Cash & Bank Balances: ₹{total_cash:,.0f}")
        st.write(f"- Market Investments: ₹{total_market_value:,.0f}")
        st.write(f"- Fixed Deposits: ₹{total_fd_value:,.0f}")
        st.write(f"- Physical Assets: ₹{total_physical_assets:,.0f}")
        if splitwise_asset > 0: st.write(f"- Splitwise Receivables: ₹{splitwise_asset:,.0f}")
    with c_li:
        st.markdown("#### 🔴 Liability Breakdown")
        st.write(f"- Credit Card Debt: ₹{total_cc_debt:,.0f}")
        st.write(f"- Active Loans: ₹{total_loan_debt:,.0f}")
        if splitwise_liability > 0: st.write(f"- Splitwise Payables: ₹{splitwise_liability:,.0f}")

    st.markdown("---")
    st.subheader("🏡 Physical Assets")
    if not assets_df.empty: st.dataframe(assets_df.style.format({'Estimated Value (₹)': '₹{:,.0f}'}), use_container_width=True, hide_index=True)
    with st.expander("➕ Register a New Physical Asset"):
        a_name = st.text_input("Asset Name (e.g., Primary Residence)")
        a_cat = st.selectbox("Category", ["Real Estate", "Vehicle", "Jewelry / Gold", "Art / Collectibles", "Other"])
        a_val = st.number_input("Current Estimated Value (₹)", min_value=0.0, step=50000.0)
        if st.button("💾 Save Asset"):
            if a_name and a_val > 0:
                try:
                    wb = openpyxl.load_workbook('Finance Tracker.xlsx')
                    if 'Physical Assets' not in wb.sheetnames: wb.create_sheet('Physical Assets'); ws = wb['Physical Assets']; ws.append(['Asset Name', 'Category', 'Estimated Value (₹)'])
                    else: ws = wb['Physical Assets']
                    ws.append([a_name, a_cat, a_val]); wb.save('Finance Tracker.xlsx'); st.success(f"✅ Added {a_name} to your assets!"); st.rerun()
                except Exception as e: st.error("Close your Excel file!")
                
    st.markdown("---")
    st.subheader("📋 Active Liabilities (Debt)")
    col_l1, col_l2 = st.columns([2, 1])
    with col_l1:
        if active_loans: st.dataframe(pd.DataFrame(active_loans).style.format({'EMI (₹)': '₹{:,.0f}', 'Outstanding Principal (₹)': '₹{:,.0f}'}), use_container_width=True, hide_index=True)
        else: st.info("No active loans tracked. Add one below!")
    with col_l2: st.markdown(f"<div class='glow-card glow-red'><div class='kpi-title'>CREDIT CARD DEBT</div><div class='kpi-value' style='color:#d50000'>₹{total_cc_debt:,.0f}</div></div>", unsafe_allow_html=True)
    
    st.markdown("---")
    
    st.markdown("### 🏛️ Official Reports")
    with st.expander("📄 Generate Ind AS Balance Sheet (Schedule III Format)"):
        st.write("This statement categorizes your personal finances into formal Current and Non-Current Assets/Liabilities per Indian Accounting Standards.")
        
        bs_data = [
            {"Particulars": "I. EQUITY AND LIABILITIES", "Amount (₹)": ""},
            {"Particulars": "1. Capital Account (Net Worth)", "Amount (₹)": f"{true_net_worth:,.2f}"},
            {"Particulars": "2. Non-Current Liabilities", "Amount (₹)": ""},
            {"Particulars": "   (a) Long-Term Borrowings", "Amount (₹)": f"{total_loan_debt:,.2f}"},
            {"Particulars": "3. Current Liabilities", "Amount (₹)": ""},
            {"Particulars": "   (a) Short-Term Borrowings (Credit Cards)", "Amount (₹)": f"{total_cc_debt:,.2f}"},
            {"Particulars": "   (b) Other Current Liabilities (Splitwise Payables)", "Amount (₹)": f"{splitwise_liability:,.2f}"},
            {"Particulars": "TOTAL EQUITY AND LIABILITIES", "Amount (₹)": f"{total_assets:,.2f}"},
            {"Particulars": "", "Amount (₹)": ""},
            {"Particulars": "II. ASSETS", "Amount (₹)": ""},
            {"Particulars": "1. Non-Current Assets", "Amount (₹)": ""},
            {"Particulars": "   (a) Property, Plant and Equipment", "Amount (₹)": f"{total_physical_assets:,.2f}"},
            {"Particulars": "2. Current Assets", "Amount (₹)": ""},
            {"Particulars": "   (a) Financial Assets - Investments", "Amount (₹)": f"{total_market_value + total_fd_value:,.2f}"},
            {"Particulars": "   (b) Cash and Cash Equivalents", "Amount (₹)": f"{total_cash:,.2f}"},
            {"Particulars": "   (c) Short-Term Loans and Advances (Splitwise Receivables)", "Amount (₹)": f"{splitwise_asset:,.2f}"},
            {"Particulars": "TOTAL ASSETS", "Amount (₹)": f"{total_assets:,.2f}"}
        ]
        bs_df = pd.DataFrame(bs_data)
        
        def format_bs(row):
            if "TOTAL" in str(row['Particulars']) or "I. EQUITY" in str(row['Particulars']) or "II. ASSETS" in str(row['Particulars']):
                return ['background-color: rgba(128,128,128,0.2); font-weight: bold'] * len(row)
            return [''] * len(row)
            
        st.dataframe(bs_df.style.apply(format_bs, axis=1), use_container_width=True, hide_index=True)
        
        csv_export = bs_df.to_csv(index=False).encode('utf-8')
        st.download_button(
            label="⬇️ Download Balance Sheet (CSV)",
            data=csv_export,
            file_name=f"Ind_AS_Balance_Sheet_{current_fy}.csv",
            mime='text/csv'
        )

elif page == "📸 AI Bill Scanner":
    st.markdown("<h2>📸 AI Receipt & Bill Scanner</h2>", unsafe_allow_html=True)
    st.write("Upload a restaurant bill, Swiggy invoice, or shopping receipt. Our AI Vision model will read the document, extract the key details, and pre-fill your ledger instantly.")
    
    with st.expander("🔑 AI Setup (Required for Real Scans)", expanded=not HAS_GENAI):
        if not HAS_GENAI: st.error("⚠️ `google-generativeai` or `Pillow` is not installed! Run `pip install google-generativeai pillow` in your terminal to enable real AI.")
        st.markdown("**Get a free API key here:** [Google AI Studio](https://aistudio.google.com/app/apikey)")
        try:
            api_key = st.secrets["GEMINI_API_KEY"]
            st.success("API Key loaded securely from secrets!")
        except Exception:
            api_key = st.text_input("Enter your Gemini API Key", type="password")

    st.markdown("<div class='scanner-box'>", unsafe_allow_html=True)
    uploaded_file = st.file_uploader("Drop your receipt image here (JPG, PNG)", type=["jpg", "jpeg", "png"])
    st.markdown("</div>", unsafe_allow_html=True)
    
    if uploaded_file is not None:
        if 'current_image_name' not in st.session_state or st.session_state['current_image_name'] != uploaded_file.name:
            st.session_state['current_image_name'] = uploaded_file.name
            for key in ['scanned_merchant', 'scanned_amount', 'scanned_date']: st.session_state.pop(key, None)

        col1, col2 = st.columns([1, 1.5])
        
        with col1:
            st.image(uploaded_file, caption="Uploaded Receipt", use_column_width=True)
            
        with col2:
            if st.button("🔍 Run AI Vision Scan", use_container_width=True):
                with st.spinner("🧠 AI is analyzing your receipt..."):
                    try:
                        if api_key and HAS_GENAI:
                            genai.configure(api_key=api_key)
                            model = genai.GenerativeModel('gemini-2.5-flash')
                            img = Image.open(uploaded_file)
                            prompt = """You are an expert financial AI data extractor. Read this receipt/invoice image carefully. Extract the following information and return ONLY a strict, valid JSON object. Do not include any conversational text, markdown formatting, or ```json backticks. CRITICAL RULES: 1. "merchant": The name of the restaurant, store, or app. Look at the logo or header. 2. "amount": Find the FINAL GRAND TOTAL. Return ONLY a number. 3. "date": Find the transaction date. Format: "YYYY-MM-DD". EXPECTED JSON FORMAT: {"merchant": "Exact Name", "amount": 0.00, "date": "2024-02-17"}"""
                            response = model.generate_content([img, prompt])
                            json_match = re.search(r'\{.*?\}', response.text, re.DOTALL)
                            if json_match:
                                data = json.loads(json_match.group(0))
                                st.session_state['scanned_merchant'] = data.get("merchant", "Unknown Merchant")
                                st.session_state['scanned_amount'] = float(data.get("amount", 0.0))
                                try: st.session_state['scanned_date'] = pd.to_datetime(data.get("date")).date()
                                except: st.session_state['scanned_date'] = datetime.datetime.today().date()
                            else: st.error("AI could not format the data correctly.")
                        else:
                            import time; time.sleep(1.5); st.warning("Running in Simulator Mode (No API Key).")
                            st.session_state['scanned_merchant'] = "Zomato / Swiggy Delivery"; st.session_state['scanned_amount'] = 845.50; st.session_state['scanned_date'] = datetime.datetime.today().date()
                    except Exception as e: st.error(f"AI Extraction Failed: {e}")
            
            if 'scanned_amount' in st.session_state:
                st.success("✅ Extraction Complete!")
                with st.form("approve_scan_form"):
                    st.write("### 📝 Review & Approve")
                    s_date = st.date_input("Transaction Date", st.session_state['scanned_date'])
                    s_merchant = st.text_input("Merchant / Details", value=st.session_state['scanned_merchant'])
                    s_cat = st.selectbox("Category", ["Dining", "Shopping", "Groceries", "Travel", "Utilities", "Other"])
                    s_amount = st.number_input("Total Amount (₹)", value=float(st.session_state['scanned_amount']))
                    existing_banks = df['Bank'].dropna().unique().tolist()
                    if "Cash / Undefined" in existing_banks: existing_banks.remove("Cash / Undefined")
                    owned_cc_names = cc_df['Card Name'].tolist() if not cc_df.empty else []
                    existing_banks = [b for b in existing_banks if b not in owned_cc_names and b != "Brokerage Account"]
                    s_bank = st.selectbox("Paid Via", ["Cash / Undefined"] + existing_banks + owned_cc_names)
                    if st.form_submit_button("💾 Approve & Save", use_container_width=True):
                        try:
                            file_path = 'Finance Tracker.xlsx'; wb = openpyxl.load_workbook(file_path); sheet = wb['Budget Tracking']; next_row = 11
                            while sheet.cell(row=next_row, column=2).value: next_row += 1
                            sheet.cell(row=next_row, column=2, value=s_date.strftime("%Y-%m-%d")); sheet.cell(row=next_row, column=3, value="Expenses"); sheet.cell(row=next_row, column=4, value=s_cat); sheet.cell(row=next_row, column=5, value=s_amount); sheet.cell(row=next_row, column=6, value=f"[Bank: {s_bank}] AI Scanned Receipt: {s_merchant}")
                            wb.save(file_path)
                            for key in ['scanned_merchant', 'scanned_amount', 'scanned_date']: st.session_state.pop(key, None)
                            st.cache_data.clear() # Clear cache again for safety
                            st.success("✅ Saved ₹{s_amount} to ledger."); st.rerun()
                        except Exception as e: st.error(f"Error saving: {e}")

elif page == "📝 Transactions":
    st.subheader(f"📝 Transactions for {selected_month if selected_month != 'All Months' else 'All Months'} {selected_fy if selected_fy != 'All Years' else 'All Years'}")
    
    edit_mode = st.toggle("✏️ Enable Edit Mode (Master Ledger)")
    if edit_mode:
        st.info("⚠️ You are editing the raw Master Ledger. Fix typos, correct amounts, or check a row on the left and hit `Delete` to remove it. Press 'Save' when done.")
        edit_df = df[['Date', 'Type', 'Category', 'Amount', 'Details']].copy()
        edit_df['Date'] = pd.to_datetime(edit_df['Date']).dt.date
        edited_master = st.data_editor(edit_df, num_rows="dynamic", use_container_width=True, column_config={"Type": st.column_config.SelectboxColumn("Type", options=["Income", "Expenses", "Savings"]), "Date": st.column_config.DateColumn("Date")})
        if st.button("💾 Save Ledger Changes"):
            try:
                wb = openpyxl.load_workbook('Finance Tracker.xlsx'); sheet = wb['Budget Tracking']
                max_r = sheet.max_row
                if max_r >= 11: sheet.delete_rows(11, max_r - 10)
                for r_idx, row in enumerate(edited_master.itertuples(), start=11):
                    sheet.cell(row=r_idx, column=2, value=str(row.Date) if pd.notnull(row.Date) else "")
                    sheet.cell(row=r_idx, column=3, value=row.Type); sheet.cell(row=r_idx, column=4, value=row.Category); sheet.cell(row=r_idx, column=5, value=row.Amount); sheet.cell(row=r_idx, column=6, value=row.Details)
                wb.save('Finance Tracker.xlsx')
                st.cache_data.clear()
                st.success("✅ Ledger updated successfully! Turn off Edit Mode to see changes."); st.rerun()
            except Exception as e: st.error(f"Error saving: {e}. Please ensure Excel is closed.")
    else:
        display_df = filtered_df.drop(columns=['Bank', 'FY', 'Year'], errors='ignore').copy()
        display_df = display_df[['Date', 'Type', 'Category', 'Amount', 'Details']].sort_values(by='Date', ascending=False)
        display_df['Date'] = display_df['Date'].dt.strftime('%Y-%m-%d')
        def color_transaction_row(row):
            if row['Type'] == 'Income': return ['background-color: rgba(0, 200, 83, 0.1); color: #00c853'] * len(row)
            elif row['Type'] == 'Expenses': return ['background-color: rgba(213, 0, 0, 0.1); color: #d50000'] * len(row)
            elif row['Type'] == 'Savings': return ['background-color: rgba(41, 98, 255, 0.1); color: #2962ff'] * len(row)
            return [''] * len(row)
        st.dataframe(display_df.style.apply(color_transaction_row, axis=1).format({'Amount': '₹{:,.2f}'}), use_container_width=True, hide_index=True)


# ==========================================
# 7. CONDITIONAL DATA ENTRY FORM
# ==========================================
if page == "🏠 Main Dashboard (I&E)":
    st.sidebar.markdown("---")
    st.sidebar.subheader("➕ Quick Add Transaction")
    new_date = st.sidebar.date_input("Date", datetime.datetime.today())
    new_type = st.sidebar.selectbox("Type", ["Income", "Expenses", "Savings"])
    
    existing_banks = df['Bank'].dropna().unique().tolist()
    if "Cash / Undefined" in existing_banks: existing_banks.remove("Cash / Undefined")
    if "Brokerage Account" in existing_banks: existing_banks.remove("Brokerage Account")
    owned_cc_names = cc_df['Card Name'].tolist() if not cc_df.empty else []
    existing_banks = [b for b in existing_banks if b not in owned_cc_names]
    
    bank_options = ["Cash / Undefined"] + existing_banks
    if owned_cc_names: bank_options += ["--- Credit Cards ---"] + owned_cc_names
    bank_options += ["+ Add New Bank Account..."]
    bank_sel = st.sidebar.selectbox("Funding Source / Bank", bank_options)
    
    if bank_sel == "+ Add New Bank Account...": actual_bank = st.sidebar.text_input("Enter Bank Name (e.g. HDFC, ICICI):")
    elif bank_sel == "--- Credit Cards ---": actual_bank = "Cash / Undefined" 
    else: actual_bank = bank_sel
    
    existing_categories = df[df['Type'] == new_type]['Category'].dropna().unique().tolist()
    if new_type == "Savings":
        for default_cat in ["Fixed Deposit", "Mutual Funds", "Stocks / ETFs", "Crypto"]:
            if default_cat not in existing_categories: existing_categories.insert(0, default_cat)
    if "+ Create New Category..." not in existing_categories: existing_categories.append("+ Create New Category...")
    
    new_category_selection = st.sidebar.selectbox("Category", existing_categories)
    if new_category_selection == "+ Create New Category...": new_category = st.sidebar.text_input("Type New Category Name:")
    else: new_category = new_category_selection
    
    extra_details = ""; auto_amount = None; sub_name = None; sub_cycle = None; is_subscription = False; is_split = False; split_with = None; who_paid = None; split_amount = 0.0; linked_goal = None

    if new_type == "Savings":
        # 🎯 GOAL LINKING ENGINE
        st.sidebar.markdown("---")
        existing_goals = goals_df['Goal Name'].unique().tolist()
        goal_option = st.sidebar.selectbox("🎯 Link to Goal (Optional)", ["None"] + existing_goals + ["+ Add New Goal..."])
        
        if goal_option == "+ Add New Goal...":
            new_goal_name = st.sidebar.text_input("Goal Name (e.g. Dream Car)")
            new_goal_target = st.sidebar.number_input("Target Amount (₹)", min_value=0.0)
            new_goal_date = st.sidebar.date_input("Target Date")
            new_goal_priority = st.sidebar.slider("Priority", 1, 10, 5)
            if new_goal_name and new_goal_target > 0:
                linked_goal = new_goal_name
        elif goal_option != "None":
            linked_goal = goal_option
            
        st.sidebar.markdown("---")
        
        if new_category == "Fixed Deposit":
            roi = st.sidebar.number_input("Rate of Interest (% p.a.)", min_value=0.0, format="%.2f")
            period = st.sidebar.number_input("Period (Months)", min_value=1, step=1)
            extra_details += f" [ROI: {roi}%, Period: {period}M, Class: Debt]"
        elif new_category in ["Stocks / ETFs", "Stocks", "Stock", "Equity"]:
            stock_option = st.sidebar.selectbox("Search Stock", ["", "➕ Type Custom..."] + nse_dropdown)
            ticker = st.sidebar.text_input("Custom Symbol") if stock_option == "➕ Type Custom..." else (re.search(r'\[([^\]]+)\]', stock_option).group(1) + ".NS" if stock_option else "")
            asset_class = st.sidebar.selectbox("Asset Class", ["Equity", "Gold", "Silver", "Debt", "Real Estate", "Other"])
            qty = st.sidebar.number_input("Quantity", min_value=0.0000, format="%.4f")
            rate = st.sidebar.number_input("Rate (₹)", min_value=0.0, format="%.2f")
            auto_amount = qty * rate; st.sidebar.info(f"**Net:** ₹{auto_amount:,.2f}")
            if ticker and qty > 0: extra_details += f" [Ticker: {ticker.upper()}, Qty: {qty}, Rate: {rate}, Class: {asset_class}]"
        elif new_category in ["Mutual Funds"]:
            selected_mf = st.sidebar.selectbox("Search MF", [""] + amfi_dropdown)
            asset_class = st.sidebar.selectbox("Asset Class", ["Equity", "Debt", "Gold", "Silver", "Hybrid", "Other"])
            qty = st.sidebar.number_input("Units", min_value=0.0000, format="%.4f")
            rate = st.sidebar.number_input("NAV (₹)", min_value=0.0, format="%.4f")
            auto_amount = qty * rate; st.sidebar.info(f"**Net:** ₹{auto_amount:,.2f}")
            if selected_mf and qty > 0: 
                ticker = re.search(r'\[(\d{6})\]', selected_mf).group(1) if re.search(r'\[(\d{6})\]', selected_mf) else ""
                extra_details += f" [Ticker: {ticker}, Qty: {qty}, Rate: {rate}, Class: {asset_class}]"
        elif new_category in ["Crypto", "Cryptocurrency"]:
            ticker = st.sidebar.text_input("Crypto Symbol", help="Use Yahoo Finance formats (e.g., BTC-USD)")
            qty = st.sidebar.number_input("Quantity (Coins)", min_value=0.000000, format="%.6f")
            rate = st.sidebar.number_input("Purchase Rate (₹)", min_value=0.0, format="%.2f")
            asset_class = "Crypto"; auto_amount = qty * rate; st.sidebar.info(f"**Net Transaction:** ₹{auto_amount:,.2f}")
            if ticker and qty > 0: extra_details += f" [Ticker: {ticker.upper()}, Qty: {qty}, Rate: {rate}, Class: {asset_class}]"

    if auto_amount is not None: new_amount = auto_amount
    else: new_amount = st.sidebar.number_input("Amount (₹)", min_value=0.0, format="%.2f")

    if new_type == "Expenses":
        st.sidebar.markdown("---")
        is_subscription = st.sidebar.checkbox("🔄 Mark as Recurring Subscription")
        if is_subscription:
            sub_name = st.sidebar.text_input("Service Name (e.g. Netflix)")
            sub_cycle = st.sidebar.selectbox("Billing Cycle", ["Monthly", "Yearly", "Quarterly"])
            
        is_split = st.sidebar.checkbox("🤝 Split this expense?")
        if is_split:
            split_opt = st.sidebar.selectbox("Split with:", split_users + ["+ Add New Person..."])
            split_with = st.sidebar.text_input("Enter Person's Name:") if split_opt == "+ Add New Person..." else split_opt
            who_paid = st.sidebar.selectbox("Who Paid?", ["Jaynik", split_with] if split_with else ["Jaynik"])
            split_amount = st.sidebar.number_input(f"Amount {split_with if who_paid == 'Jaynik' else 'Jaynik'} Owes (₹)", min_value=0.0, value=float(new_amount/2) if new_amount else 0.0)

    new_notes = st.sidebar.text_input("Details / Notes (Optional)")

    if actual_bank and actual_bank != "Cash / Undefined": extra_details += f" [Bank: {actual_bank}]"
    if is_subscription and sub_name: extra_details += f" [Sub: {sub_name}]"
    if linked_goal: extra_details += f" [Goal: {linked_goal}]"
    final_details = (new_notes + extra_details).strip()

    if st.sidebar.button("💾 Save Transaction", use_container_width=True):
        if new_category == "" or new_amount == 0: 
            st.sidebar.error("⚠️ Please enter a Category and Amount above 0.")
        else:
            try:
                file_path = 'Finance Tracker.xlsx'
                wb = openpyxl.load_workbook(file_path)
                
                # 1. Save Goal if New
                if linked_goal and goal_option == "+ Add New Goal...":
                    if 'Goals' not in wb.sheetnames:
                        ws_goals = wb.create_sheet('Goals')
                        ws_goals.append(['Goal Name', 'Target Amount', 'Target Date', 'Priority', 'Status'])
                    else: ws_goals = wb['Goals']
                    ws_goals.append([new_goal_name, new_goal_target, new_goal_date.strftime("%Y-%m-%d"), new_goal_priority, "Active"])

                # 2. Save to Main Ledger
                sheet = wb['Budget Tracking']
                next_row = 11
                while sheet.cell(row=next_row, column=2).value: next_row += 1
                sheet.cell(row=next_row, column=2, value=new_date.strftime("%Y-%m-%d"))
                sheet.cell(row=next_row, column=3, value=new_type)
                sheet.cell(row=next_row, column=4, value=new_category)
                sheet.cell(row=next_row, column=5, value=new_amount)
                sheet.cell(row=next_row, column=6, value=final_details)
                
                # 3. Automatically push to Subscription Radar
                if new_type == "Expenses" and is_subscription and sub_name:
                    if 'Subscriptions' not in wb.sheetnames:
                        ws_sub = wb.create_sheet('Subscriptions')
                        ws_sub.append(['Service Name', 'Category', 'Billing Cycle', 'Amount (₹)', 'Next Due Date'])
                    else: ws_sub = wb['Subscriptions']
                    base_date = pd.to_datetime(new_date)
                    if sub_cycle == 'Monthly': next_due = base_date + pd.DateOffset(months=1)
                    elif sub_cycle == 'Yearly': next_due = base_date + pd.DateOffset(years=1)
                    else: next_due = base_date + pd.DateOffset(months=3)
                    ws_sub.append([sub_name, new_category, sub_cycle, new_amount, next_due.strftime("%Y-%m-%d")])
                    
                # 4. Automatically push to Splitwise
                if new_type == "Expenses" and is_split and split_with:
                    if 'Splitwise' not in wb.sheetnames:
                        ws_split = wb.create_sheet('Splitwise')
                        ws_split.append(['ID', 'Date', 'Payer', 'Debtor', 'Total Amount', 'Split Amount', 'Description', 'Status'])
                    else: ws_split = wb['Splitwise']
                    split_id = f"SPL-{datetime.datetime.now().strftime('%Y%m%d%H%M%S')}"
                    debtor = split_with if who_paid == "Jaynik" else "Jaynik"
                    sp_desc = new_notes if new_notes else new_category
                    ws_split.append([split_id, new_date.strftime("%Y-%m-%d"), who_paid, debtor, new_amount, split_amount, sp_desc, "Pending"])
                
                wb.save(file_path)
                st.cache_data.clear() # IMPORTANT: Clears cache so new data shows up instantly
                st.sidebar.success(f"✅ Saved completely!")
                st.rerun()
            except Exception as e: 

                st.sidebar.error(f"Error saving data: {e}. Is Excel open?")







